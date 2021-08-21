from os import path, makedirs
import extract_data, constants as const

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter, exceptions as excel_exceptions


def export_info_to_excel_format(ptu_obj: extract_data.PtuWorkBook, ptu_workbook: Workbook) -> None:
    """
    The main function for exporting ptu object data into excel format.
    :param ptu_obj: input, containing information of ptu object
    :param ptu_workbook: output, excel workbook that can be saved later
    :return: None
    """

    class Coordinate:
        """
        Class for representing coordinate of an excel workbook cell.
        """

        def __init__(self, y, x):
            self.row = y
            self.column = x

    def write_cell_info(cell, value, horizontal_alignment="left", font_name="Calibri", font_size=11, is_bold=False,
                        is_wrapped_text=False, style="Normal", is_italic=False):
        """
        This function is used almost in every other functions for writing info.
        Input is cell features and value will be written into cell according to features.
        Some default features are also considered.
        """
        try:
            cell.value = value
        except excel_exceptions.IllegalCharacterError:  # When value string contains a character which is not writeable
            print(value)
            cell.value = "IllegalCharacterError while parsing"
            is_italic = True
        cell.style = style
        cell.font = Font(name=font_name, size=font_size, bold=is_bold, italic=is_italic)
        cell.alignment = Alignment(horizontal=horizontal_alignment, vertical="center", wrap_text=is_wrapped_text)

    def get_coordinate(defined_name):
        """
        There are some cells in excel workbooks that can be defined by a name.
        This function returns coordinate of a cell with given defined name and also name of related workbook.
        """
        defined_names = ptu_workbook.defined_names
        for title, coord in defined_names[defined_name].destinations:
            cell = ptu_workbook[title][coord]
            coordinate = Coordinate(cell.row, cell.column)
            return title, coordinate

    def get_cell(cell_name):
        """
        :returns cell of excel workbook with defined name "cell_name"
        """
        worksheet_title, coord = get_coordinate(cell_name)
        return ptu_workbook[worksheet_title].cell(coord.row, coord.column)

    def write_preface_info():
        """
         Create and write information of preface worksheet
        """
        write_cell_info(get_cell("Purpose"), value=ptu_obj.preface.purpose)
        write_cell_info(get_cell("Processor"), value=ptu_obj.preface.processor)
        write_cell_info(get_cell("Tool_chain"), value=ptu_obj.preface.tool_chain)
        write_cell_info(get_cell("HEADER.module_name"), value=ptu_obj.preface.header.module_name,
                        horizontal_alignment="center")
        write_cell_info(get_cell("HEADER.module_version"), value=ptu_obj.preface.header.module_version,
                        horizontal_alignment="center")
        write_cell_info(get_cell("HEADER.test_plan_version"), value=ptu_obj.preface.header.test_plan_version,
                        horizontal_alignment="center")

    def write_include_info():
        """
         Write information of included header files in include worksheet
        """
        worksheet_title, coord = get_coordinate("include")
        worksheet = ptu_workbook[worksheet_title]
        for counter in range(len(ptu_obj.include)):
            cell = worksheet.cell(coord.row + counter, coord.column)
            write_cell_info(cell, value=ptu_obj.include[counter])

    def write_comment_info():
        """
         Write comments in COMMENT worksheet
        """
        worksheet_title, coord = get_coordinate("COMMENT")
        worksheet = ptu_workbook[worksheet_title]
        for counter in range(len(ptu_obj.comment)):
            cell = worksheet.cell(coord.row + counter, coord.column)
            write_cell_info(cell, value=ptu_obj.comment[counter])

    def write_user_code_info():
        """

        """
        worksheet_title, coord = get_coordinate("USER_CODE")
        worksheet = ptu_workbook[worksheet_title]

        counter = 0
        if ptu_obj.user_code:
            cell = worksheet.cell(coord.row, coord.column)
            write_cell_info(cell, value="Before Services", horizontal_alignment="center", style="Input", is_bold=True)

            counter += 1
            user_code_start_row = coord.row + counter
            for user_code in ptu_obj.user_code:
                cell = worksheet.cell(coord.row + counter, coord.column)
                write_cell_info(cell, value=user_code.code)

                cell = worksheet.cell(coord.row + counter, coord.column - 1)
                write_conditions(cell, user_code.conditions)

                counter += 1
            user_code_end_row = coord.row + counter - 1
            worksheet.row_dimensions.group(user_code_start_row, user_code_end_row, hidden=True)

        services_sheet_title, services_sheet_row_num = get_coordinate("SERVICE")
        services_sheet_row_num = services_sheet_row_num.row - 1

        for service in ptu_obj.services:
            services_sheet_row_num += 1
            if not service.has_user_code():
                continue

            cell = worksheet.cell(coord.row + counter, coord.column)
            write_cell_info(cell, value="In Service \"" + service.name + "\"", is_bold=True, style="Input",
                            horizontal_alignment="center")
            cell.hyperlink = "#" + services_sheet_title + "!A" + str(services_sheet_row_num)
            ptu_workbook[services_sheet_title]["A" + str(services_sheet_row_num)].hyperlink = \
                "#" + worksheet.title + "!B" + str(cell.row)

            counter += 1
            service_user_code_start_row = coord.row + counter
            if service.user_code:
                cell = worksheet.cell(coord.row + counter, coord.column)
                write_cell_info(cell, value="Common", style="Output", horizontal_alignment="center")

                counter += 1
                for user_code in service.user_code:
                    cell = worksheet.cell(coord.row + counter, coord.column)
                    write_cell_info(cell, value=user_code.code)

                    cell = worksheet.cell(coord.row + counter, coord.column - 1)
                    write_conditions(cell, user_code.conditions)

                    counter += 1

            for test in service.test_list:
                services_sheet_row_num += 1
                element = test.element
                if not element.user_code:
                    continue

                cell = worksheet.cell(coord.row + counter, coord.column)
                write_cell_info(cell, value="In Test \"" + test.name + "\"", style="Output",
                                horizontal_alignment="center")
                cell.hyperlink = "#TEST_ROW" + str(services_sheet_row_num) + "!A1"

                test_worksheet = ptu_workbook["TEST_ROW" + str(services_sheet_row_num)]
                test_worksheet["A" + str(test_worksheet.max_row - 1)].hyperlink = \
                    "#" + worksheet.title + "!B" + str(cell.row)

                counter += 1
                for user_code in element.user_code:
                    cell = worksheet.cell(coord.row + counter, coord.column)
                    write_cell_info(cell, value=user_code.code)

                    cell = worksheet.cell(coord.row + counter, coord.column - 1)
                    write_conditions(cell, user_code.conditions)

                    counter += 1
            service_user_code_end_row = coord.row + counter - 1
            worksheet.row_dimensions.group(service_user_code_start_row, service_user_code_end_row, hidden=True)

    def write_conditions(cell, conditions, style="Normal"):
        """
        This function is used for writing conditions of any type of data in the given cell
        """
        all_conditions = ",".join(conditions)  # Concatenate all conditions
        write_cell_info(cell, value=all_conditions, style=style)

    def write_stub_definitions_info():
        """
        Writes all stub definitions and their content in STUBS sheet
        """
        worksheet_title, coord = get_coordinate("DEFINE_STUB")
        worksheet = ptu_workbook[worksheet_title]
        row_counter = 0
        for define_stub in ptu_obj.stub_definitions:
            # Writing define stub name and conditions
            cell = worksheet.cell(coord.row + row_counter, coord.column)
            write_cell_info(cell, value=define_stub.name, horizontal_alignment="center", is_bold=True, style="Output")

            cell = worksheet.cell(coord.row + row_counter, coord.column + 1)
            write_conditions(cell, define_stub.conditions, style="Output")

            # Writing Stubs
            row_counter += 1
            define_stub_row_start = coord.row + row_counter
            for stub in define_stub.stub_list:
                cell = worksheet.cell(coord.row + row_counter, coord.column)
                write_cell_info(cell, value=stub.stub_definition, is_wrapped_text=True)

                cell = worksheet.cell(coord.row + row_counter, coord.column + 1)
                write_conditions(cell, stub.conditions)

                row_counter += 1

            # Grouping Stub Definitions
            define_stub_row_end = coord.row + row_counter - 1
            worksheet.row_dimensions.group(define_stub_row_start, define_stub_row_end, hidden=True)

    def write_initialisation_info():
        """
        Initialisation in ptu files is optional.
        If there is an initialisation scope in ptu file, this function writes it in INITIALISATION sheet.
        """
        worksheet_title, coord = get_coordinate("INITIALISATION")
        worksheet = ptu_workbook[worksheet_title]

        counter = 0
        if ptu_obj.initialisation:
            for initialisation in ptu_obj.initialisation:
                cell = worksheet.cell(coord.row + counter, coord.column)
                write_cell_info(cell, value=initialisation.description)

                cell = worksheet.cell(coord.row + counter, coord.column + 1)
                write_conditions(cell, initialisation.conditions)

                counter += 1

    def write_test_case_info(worksheet, coord, test_case):
        """
        Any test has multiple parameters as test cases.
        For each of test cases in a single test, this function is called.
        """
        cell = worksheet.cell(coord.row, coord.column)
        write_cell_info(cell, value=test_case.param_type)

        cell = worksheet.cell(coord.row, coord.column + 1)
        write_cell_info(cell, value=test_case.param_name)

        cell = worksheet.cell(coord.row, coord.column + 2)
        write_cell_info(cell, value=test_case.init)

        cell = worksheet.cell(coord.row, coord.column + 3)
        write_cell_info(cell, value=test_case.ev)

    def write_environments_info():
        """
        Environments in ptu files are optional.
        This function writes all environments and their test cases in ENVIRONMENT worksheet
        """
        worksheet_title, coord = get_coordinate("ENVIRONMENT")
        worksheet = ptu_workbook[worksheet_title]
        row_counter = 0
        for environment in ptu_obj.environments:
            # Writing Environment name and conditions
            cell = worksheet.cell(coord.row + row_counter, coord.column)
            write_cell_info(cell, value=environment.name, horizontal_alignment="center", is_bold=True, style="Output")
            worksheet.merge_cells(start_row=cell.row, start_column=cell.column, end_row=cell.row,
                                  end_column=cell.column + const.ENVIRONMENT_WIDTH - 1)

            cell = worksheet.cell(coord.row + row_counter, coord.column + const.ENVIRONMENT_WIDTH)
            write_conditions(cell, environment.conditions, style="Output")

            # Writing test cases
            row_counter += 1
            environment_row_start = coord.row + row_counter
            for test_case in environment.test_case_list:
                write_test_case_info(worksheet, Coordinate(coord.row + row_counter, coord.column), test_case)

                cell = worksheet.cell(coord.row + row_counter, coord.column + const.ENVIRONMENT_WIDTH)
                write_conditions(cell, test_case.conditions)

                row_counter += 1

            # Grouping Environments
            environment_row_end = coord.row + row_counter - 1
            worksheet.row_dimensions.group(environment_row_start, environment_row_end, hidden=True)

    def write_test_info(test, row, service_name):
        """
        Every service has multiple tests. Data in each of them should be written in a new sheet.
        This function writes info of a single test in a new sheet. It is called in a loop in write_service_info func.
        :param test: Info of this test is written in a new sheet.
        :param row: Row number of test in service sheet. Used for naming test sheet and also for hyperlink.
        :param service_name: Service name is shown at the top row of the sheet.
        """
        # Create new worksheet
        worksheet = ptu_workbook.create_sheet(title="TEST_ROW" + str(row))
        coord = Coordinate(1, 1)
        row_counter = 1

        """ Writing Title Part"""

        cell = worksheet.cell(coord.row, coord.column)
        title = "SERVICE NAME : " + service_name
        title += "                                       "
        title += "==================================="
        title += "                                       "
        title += "TEST NAME : " + test.name
        write_cell_info(cell, value=title, style="Check Cell", font_size=14, is_bold=True,
                        horizontal_alignment="center")
        worksheet.merge_cells(start_row=cell.row, start_column=cell.column, end_row=cell.row,
                              end_column=cell.column + const.ELEMENT_WIDTH - 1)
        # Freezing top row of sheet
        worksheet.freeze_panes = 'A2'

        """ Writing Comment part """

        cell = worksheet.cell(coord.row + row_counter, coord.column)
        write_cell_info(cell, value="COMMENTS", horizontal_alignment="center", style="Accent6")
        worksheet.merge_cells(start_row=cell.row, start_column=cell.column, end_row=cell.row,
                              end_column=cell.column + const.ELEMENT_WIDTH - 1)

        row_counter += 1
        comment_row_start = coord.row + row_counter

        # If there was no COMMENT, write an empty row
        if not test.comment:
            test.comment.append("")
        # Writing all comments
        for comment in test.comment:
            cell = worksheet.cell(coord.row + row_counter, coord.column)
            write_cell_info(cell, value=comment)
            worksheet.merge_cells(start_row=cell.row, start_column=cell.column, end_row=cell.row,
                                  end_column=cell.column + const.ELEMENT_WIDTH - 1)
            row_counter += 1

        comment_row_end = coord.row + row_counter - 1
        # Grouping Comments
        worksheet.row_dimensions.group(comment_row_start, comment_row_end, hidden=True)

        # Getting element of test
        element = test.element

        """ Writing USE part """

        cell = worksheet.cell(coord.row + row_counter, coord.column)
        write_cell_info(cell, value="USE", horizontal_alignment="center", style="Accent2")
        worksheet.merge_cells(start_row=cell.row, start_column=cell.column, end_row=cell.row,
                              end_column=cell.column + const.ELEMENT_WIDTH - 1)

        row_counter += 1
        use_row_start = coord.row + row_counter  # !

        # If there was no USE, write an empty row
        if not element.use:
            element.use.append("")
        # Writing all uses
        for use in element.use:
            cell = worksheet.cell(coord.row + row_counter, coord.column)
            write_cell_info(cell, value=use)
            worksheet.merge_cells(start_row=cell.row, start_column=cell.column, end_row=cell.row,
                                  end_column=cell.column + const.ELEMENT_WIDTH - 1)
            row_counter += 1

        use_row_end = coord.row + row_counter - 1
        # Grouping uses
        worksheet.row_dimensions.group(use_row_start, use_row_end, hidden=True)

        """ Writing TEST DATA PART"""

        cell = worksheet.cell(coord.row + row_counter, coord.column)
        write_cell_info(cell, value="TEST DATA", horizontal_alignment="center", style="Accent1")
        worksheet.merge_cells(start_row=cell.row, start_column=cell.column, end_row=cell.row,
                              end_column=cell.column + const.ELEMENT_WIDTH - 1)
        row_counter += 1

        """ Writing TEST DATA PART - IDENTIFIER COLUMNS """

        test_data_columns = ["Conditions", "<param>", "<name>", "init", "ev"]
        for column_counter in range(const.ELEMENT_WIDTH):
            cell = worksheet.cell(coord.row + row_counter, coord.column + column_counter)
            write_cell_info(cell, value=test_data_columns[column_counter], font_size=14, is_bold=True,
                            horizontal_alignment="center")
            # Setting default column dimensions for worksheet
            worksheet.column_dimensions[get_column_letter(column_counter + 1)].width = \
                const.TEST_SHEET_COLUMNS_WIDTH[column_counter]
        row_counter += 1

        """ Writing TEST DATA - input data"""

        cell = worksheet.cell(coord.row + row_counter, coord.column)
        write_cell_info(cell, value="input data", horizontal_alignment="center", style="Output")
        worksheet.merge_cells(start_row=cell.row, start_column=cell.column, end_row=cell.row,
                              end_column=cell.column + const.ELEMENT_WIDTH - 1)
        row_counter += 1
        input_data_row_start = coord.row + row_counter

        # If there was no input data, write an empty row
        if not element.input_data:
            element.input_data.append(extract_data.PtuWorkBook.TestCase())
        # Writing all input data
        for test_case in element.input_data:
            cell = worksheet.cell(coord.row + row_counter, coord.column)
            write_conditions(cell, test_case.conditions)

            cell = worksheet.cell(coord.row + row_counter, coord.column + 1)
            write_test_case_info(worksheet, Coordinate(cell.row, cell.column), test_case)

            row_counter += 1

        input_data_row_end = coord.row + row_counter - 1
        # Grouping all input data
        worksheet.row_dimensions.group(input_data_row_start, input_data_row_end, hidden=True)

        """ Writing TEST DATA - calibrations"""

        cell = worksheet.cell(coord.row + row_counter, coord.column)
        write_cell_info(cell, value="calibrations", horizontal_alignment="center", style="Output")
        worksheet.merge_cells(start_row=cell.row, start_column=cell.column, end_row=cell.row,
                              end_column=cell.column + const.ELEMENT_WIDTH - 1)
        row_counter += 1
        calibrations_row_start = coord.row + row_counter

        # If there was no calibrations, write an empty row
        if not element.calibrations:
            element.calibrations.append(extract_data.PtuWorkBook.TestCase())
        # Writing all calibrations
        for test_case in element.calibrations:
            cell = worksheet.cell(coord.row + row_counter, coord.column)
            write_conditions(cell, test_case.conditions)

            cell = worksheet.cell(coord.row + row_counter, coord.column + 1)
            write_test_case_info(worksheet, Coordinate(cell.row, cell.column), test_case)

            row_counter += 1

        calibrations_row_end = coord.row + row_counter - 1
        # Grouping all calibrations
        worksheet.row_dimensions.group(calibrations_row_start, calibrations_row_end, hidden=True)

        """ Writing TEST DATA - output data """

        cell = worksheet.cell(coord.row + row_counter, coord.column)
        write_cell_info(cell, value="output data", horizontal_alignment="center", style="Output")
        worksheet.merge_cells(start_row=cell.row, start_column=cell.column, end_row=cell.row,
                              end_column=cell.column + const.ELEMENT_WIDTH - 1)
        row_counter += 1
        output_data_row_start = coord.row + row_counter

        # If there was no output data, write an empty row
        if not element.output_data:
            element.output_data.append(extract_data.PtuWorkBook.TestCase())
        # Writing all output data
        for test_case in element.output_data:
            cell = worksheet.cell(coord.row + row_counter, coord.column)
            write_conditions(cell, test_case.conditions)

            cell = worksheet.cell(coord.row + row_counter, coord.column + 1)
            write_test_case_info(worksheet, Coordinate(cell.row, cell.column), test_case)

            row_counter += 1

        output_data_row_end = coord.row + row_counter - 1
        # Grouping all output data
        worksheet.row_dimensions.group(output_data_row_start, output_data_row_end, hidden=True)

        """ Writing STUB part """

        cell = worksheet.cell(coord.row + row_counter, coord.column)
        write_cell_info(cell, value="STUB", horizontal_alignment="center", style="Accent5")
        worksheet.merge_cells(start_row=cell.row, start_column=cell.column, end_row=cell.row,
                              end_column=cell.column + const.ELEMENT_WIDTH - 1)
        row_counter += 1

        """ Writing STUB part - IDENTIFIER COLUMNS """

        cell = worksheet.cell(coord.row + row_counter, coord.column)
        write_cell_info(cell, value="Conditions", font_size=14, is_bold=True, horizontal_alignment="center")

        cell = worksheet.cell(coord.row + row_counter, coord.column + 1)
        write_cell_info(cell, value="<stub>", font_size=14, is_bold=True, horizontal_alignment="center")
        worksheet.merge_cells(start_row=cell.row, start_column=cell.column, end_row=cell.row,
                              end_column=cell.column + const.ELEMENT_WIDTH - 2)

        row_counter += 1
        stub_row_start = coord.row + row_counter

        """ Writing STUB part - STUBs """

        # If there was no stub call data, write an empty row
        if not element.stub_calls:
            element.stub_calls.append(extract_data.PtuWorkBook.Stub())
        # Writing all STUB
        for stub in element.stub_calls:
            cell = worksheet.cell(coord.row + row_counter, coord.column)
            write_conditions(cell, stub.conditions)

            cell = worksheet.cell(coord.row + row_counter, coord.column + 1)
            write_cell_info(cell, value=stub.stub_definition)
            worksheet.merge_cells(start_row=cell.row, start_column=cell.column, end_row=cell.row,
                                  end_column=cell.column + const.ELEMENT_WIDTH - 2)

            row_counter += 1

        stub_row_end = coord.row + row_counter - 1
        # Grouping all Stubs
        worksheet.row_dimensions.group(stub_row_start, stub_row_end, hidden=True)

        """ Writing USER CODE part - A row linked to user code sheet"""

        cell = worksheet.cell(coord.row + row_counter, coord.column)
        write_cell_info(cell, value="See User Code for this test", style="40 % - Accent4", is_italic=True,
                        horizontal_alignment="center")
        worksheet.merge_cells(start_row=cell.row, start_column=cell.column, end_row=cell.row,
                              end_column=cell.column + const.ELEMENT_WIDTH - 1)
        row_counter += 1

        """ A row for returning to services sheet"""

        cell = worksheet.cell(coord.row + row_counter, coord.column)
        write_cell_info(cell, value="Return to Service list", horizontal_alignment="center", style="Bad",
                        is_italic=True)
        worksheet.merge_cells(start_row=cell.row, start_column=cell.column, end_row=cell.row,
                              end_column=cell.column + const.ELEMENT_WIDTH - 1)
        cell.hyperlink = "#SERVICES!A" + str(row)

        return worksheet

    def write_services_info():
        """
        A single ptu file has multiple services(functions) to test.
        For each of them, this function is called to write info of that service in SERVICES sheet.
        """
        worksheet_title, coord = get_coordinate("SERVICE")
        worksheet = ptu_workbook[worksheet_title]
        row_counter = 0
        for service in ptu_obj.services:
            # Writing service name and conditions
            cell = worksheet.cell(coord.row + row_counter, coord.column)
            write_cell_info(cell, value=service.name, horizontal_alignment="center", is_bold=True, style="Output")
            worksheet.merge_cells(start_row=cell.row, start_column=cell.column, end_row=cell.row,
                                  end_column=cell.column + const.SERVICE_WIDTH - 1)

            cell = worksheet.cell(coord.row + row_counter, coord.column + const.SERVICE_WIDTH)
            write_conditions(cell, service.conditions, style="Output")

            # Writing test cases
            row_counter += 1
            service_row_start = coord.row + row_counter
            for test in service.test_list:
                # Creating test sheet
                new_worksheet = write_test_info(test, coord.row + row_counter, service.name)

                cell = worksheet.cell(coord.row + row_counter, coord.column)
                write_cell_info(cell, value=test.name)
                cell.hyperlink = "#" + new_worksheet.title + "!A1"

                cell = worksheet.cell(coord.row + row_counter, coord.column + 1)
                write_cell_info(cell, value=test.family)

                cell = worksheet.cell(coord.row + row_counter, coord.column + const.SERVICE_WIDTH)
                write_conditions(cell, test.conditions)

                row_counter += 1

            # Grouping services
            service_row_end = coord.row + row_counter - 1
            worksheet.row_dimensions.group(service_row_start, service_row_end, hidden=True)

    """ Calling all of functions to write their own part in the workbook """

    write_preface_info()
    write_include_info()
    write_comment_info()
    write_stub_definitions_info()
    write_initialisation_info()
    write_environments_info()
    write_services_info()
    write_user_code_info()


def save_workbook(workbook, name, output_path):
    """
    Just saves workbook in output_path with the given name
    """
    if not path.exists(output_path):
        makedirs(output_path)  # If folder doesn't exist, create it
        print("Output path didn't exist")
    workbook.save(output_path + "\\" + name + ".xlsx")
